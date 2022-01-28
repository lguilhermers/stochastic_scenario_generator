"""
This tool automates a number of parallelized simulations based on variable distributions, saving main inputs and outputs
"""
from __future__ import division
from __future__ import print_function

import cea.config
import cea.inputlocator
import numpy as np
import pandas as pd
import os
from cea.datamanagement import archetypes_mapper
from cea.demand import demand_main, schedule_maker
from cea.demand.schedule_maker import schedule_maker
from cea.resources.radiation_daysim import radiation_main
from openpyxl.reader.excel import load_workbook
from geopandas import GeoDataFrame as Gdf
from cea.utilities.dbf import dbf_to_dataframe, dataframe_to_dbf
from scipy.stats import beta
import random
import shutil
import multiprocessing
from itertools import repeat
import cea.plugin

__author__ = "Luis Santos"
__copyright__ = "Copyright 2020, Architecture and Building Systems - ETH Zurich"
__credits__ = ["Luis Santos, Shanshan Hsieh, Reynold Mok"]
__license__ = "MIT"
__version__ = "1.0"
__maintainer__ = "Shanshan Hsieh"
__email__ = "cea@arch.ethz.ch"
__status__ = "Production"

class ScenarioPlugin(cea.plugin.CeaPlugin):
    pass

def stochastic_scenario_generator(config, locator, dataframe_with_instances):
    """
    This function loads inputs from a dataframe into the CEA database, runs CEA scripts and stores inputs and outputs
    """

    outputs_list = []
    # loop inside input dataframe and replaces database for every iteration
    for index, instance in dataframe_with_instances.iterrows():
        # config.multiprocessing = True  # assures each simulation uses a single core
        print("Simulation number " + str(index))

        # Replace STANDARDs to account for the correct database
        typology = dbf_to_dataframe(locator.get_building_typology())
        typology.STANDARD = 'STANDARD1'
        dataframe_to_dbf(typology, locator.get_building_typology())

        types = load_workbook(filename=locator.get_database_construction_standards())
        envelope_types = types['ENVELOPE_ASSEMBLIES']
        envelope_types.cell(column=2, row=2).value = 'CONSTRUCTION_AS1'
        envelope_types.cell(column=3, row=2).value = 'TIGHTNESS_AS1'
        envelope_types.cell(column=4, row=2).value = 'WINDOW_AS1'
        envelope_types.cell(column=5, row=2).value = 'ROOF_AS1'
        envelope_types.cell(column=6, row=2).value = 'WALL_AS1'
        envelope_types.cell(column=7, row=2).value = 'WALL_AS1'
        envelope_types.cell(column=10, row=2).value = 'SHADING_AS0'
        hvac_types = types['HVAC_ASSEMBLIES']
        hvac_types.cell(column=3, row=2).value = 'HVAC_COOLING_AS3'  # baseline: centralized
        hvac_types.cell(column=4, row=2).value = 'HVAC_HOTWATER_AS4'  # baseline: high temperature in tropics
        hvac_types.cell(column=5, row=2).value = 'HVAC_CONTROLLER_AS2'  # baseline: PI controller
        hvac_types.cell(column=6, row=2).value = 'HVAC_VENTILATION_AS1'  # baseline: mechanical ventilation
        supply_types = types['SUPPLY_ASSEMBLIES']
        supply_types.cell(column=3, row=2).value = 'SUPPLY_HOTWATER_AS1'  # baseline: electric boiler
        supply_types.cell(column=4, row=2).value = 'SUPPLY_COOLING_AS1'  # baseline: VCC and dry CT
        types.save(locator.get_database_construction_standards())

        # Replace database for inputs generated

        # Changes and saves variables related to ZONE
        zone_gdf = Gdf.from_file(locator.get_zone_geometry())
        zone_gdf['floors_ag'] = instance.zone_floors_ag
        zone_gdf['height_ag'] = instance.zone_floor_to_floor_height * zone_gdf['floors_ag']
        zone_gdf['GFA_m2'] = zone_gdf.area * (zone_gdf['floors_ag'] + zone_gdf['floors_bg'])
        zone_gdf.to_file(locator.get_zone_geometry())

        # # Changes and saves variables related to SURROUNDINGS
        surroundings_gdf = Gdf.from_file(locator.get_surroundings_geometry())
        surroundings_gdf.floors_ag = [instance.surrounding_floors_ag for floor in surroundings_gdf.floors_ag]
        surroundings_height = instance.surroundings_floor_to_floor_height * instance.surrounding_floors_ag
        surroundings_gdf['height_ag'] = [surroundings_height for height in surroundings_gdf.height_ag]
        surroundings_gdf.to_file(locator.get_surroundings_geometry())

        # Changes and saves variables related to CONSTRUCTION_STANDARDS
        archetype_construction = load_workbook(filename=locator.get_database_construction_standards())
        envelope_assemblies = archetype_construction['ENVELOPE_ASSEMBLIES']
        envelope_assemblies.cell(column=11, row=2).value = instance.Es
        envelope_assemblies.cell(column=12, row=2).value = instance.Hs_ag
        envelope_assemblies.cell(column=14, row=2).value = instance.Ns
        envelope_assemblies.cell(column=15, row=2).value = instance.void_deck
        envelope_assemblies.cell(column=16, row=2).value = instance.wwr
        envelope_assemblies.cell(column=17, row=2).value = instance.wwr
        envelope_assemblies.cell(column=18, row=2).value = instance.wwr
        envelope_assemblies.cell(column=19, row=2).value = instance.wwr
        archetype_construction.save(locator.get_database_construction_standards())

        # Changes and saves variables related to ENVELOPE
        assemblies_envelope = load_workbook(filename=locator.get_database_envelope_systems())
        construction = assemblies_envelope['CONSTRUCTION']
        construction.cell(column=3, row=2).value = instance.Cm_Af
        tightness = assemblies_envelope['TIGHTNESS']
        tightness.cell(column=3, row=2).value = instance.n50
        window = assemblies_envelope['WINDOW']
        window.cell(column=3, row=2).value = instance.U_win
        window.cell(column=4, row=2).value = instance.G_win
        window.cell(column=5, row=2).value = instance.e_win
        window.cell(column=6, row=2).value = instance.F_F
        roof = assemblies_envelope['ROOF']
        roof.cell(column=3, row=2).value = instance.U_roof
        roof.cell(column=4, row=2).value = instance.a_roof
        roof.cell(column=5, row=2).value = instance.e_roof
        roof.cell(column=6, row=2).value = instance.r_roof
        wall = assemblies_envelope['WALL']
        wall.cell(column=3, row=2).value = instance.U_wall
        wall.cell(column=4, row=2).value = instance.a_wall
        wall.cell(column=5, row=2).value = instance.e_wall
        wall.cell(column=6, row=2).value = instance.r_wall
        shading = assemblies_envelope['SHADING']
        shading.cell(column=3, row=2).value = instance.rf_sh
        assemblies_envelope.save(locator.get_database_envelope_systems())

        # Changes and saves variables related to USE_TYPE_PROPERTIES
        archetypes_use_type = load_workbook(filename=locator.get_database_use_types_properties())
        internal_loads = archetypes_use_type['INTERNAL_LOADS']
        internal_loads.cell(column=2, row=5).value = instance.Occ_m2p
        internal_loads.cell(column=3, row=5).value = instance.Qs_Wp
        internal_loads.cell(column=4, row=5).value = instance.X_ghp
        internal_loads.cell(column=5, row=5).value = instance.Ea_Wm2
        internal_loads.cell(column=6, row=5).value = instance.El_Wm2
        internal_loads.cell(column=9, row=5).value = instance.Vww_ldp
        indoor_comfort = archetypes_use_type['INDOOR_COMFORT']
        indoor_comfort.cell(column=2, row=5).value = instance.Tcs_set_C
        indoor_comfort.cell(column=6, row=5).value = instance.Ve_lsp
        archetypes_use_type.save(locator.get_database_use_types_properties())

        # Changes and saves variables related to HVAC
        archetypes_use_type = load_workbook(filename=locator.get_database_air_conditioning_systems())
        internal_loads = archetypes_use_type['CONTROLLER']
        internal_loads.cell(column=4, row=4).value = instance.dT_Qcs
        internal_loads = archetypes_use_type['VENTILATION']
        internal_loads.cell(column=5, row=3).value = instance.HEAT_REC
        internal_loads.cell(column=7, row=3).value = instance.ECONOMIZER
        indoor_comfort = archetypes_use_type['COOLING']
        indoor_comfort.cell(column=4, row=5).value = instance.convection_cs
        indoor_comfort.cell(column=6, row=5).value = instance.dTcs_C
        archetypes_use_type.save(locator.get_database_air_conditioning_systems())

        # Changes and saves variables related to SUPPLY
        archetypes_use_type = load_workbook(filename=locator.get_database_supply_assemblies())
        internal_loads = archetypes_use_type['COOLING']
        internal_loads.cell(column=6, row=3).value = instance.efficiency_cooling
        archetypes_use_type.save(locator.get_database_supply_assemblies())

        # Run CEA scripts: archetypes, solar radiation, building schedules and energy demand

        config.multiprocessing = False  # assures each simulation uses a single core
        config.debug = False
        config.scenario = locator.scenario
        # config.scenario_name = config.scenario.split('\\')[-1]
        config.scenario_name = config.scenario.rsplit(os.sep)[-1]
        archetypes_mapper.main(config)  # loads database into the scenario
        radiation_main.main(config)  # runs solar radiation script
        schedule_maker.schedule_maker_main(locator, config)  # runs schedules
        demand_main.demand_calculation(locator, config)  # runs demand simulation
        # config.multiprocessing = True

        # Process relevant outputs

        # Total weekly schedules are calculated (calculated by the weekly_schedule function)
        schedules = pd.read_csv(locator.get_building_weekly_schedules('B1001'), skiprows=2)
        weekly_occupancy_h = round(weekly_schedule(schedules, 'OCCUPANCY'), 1)
        weekly_appliances_h = round(weekly_schedule(schedules, 'APPLIANCES'), 1)
        weekly_lighting_h = round(weekly_schedule(schedules, 'LIGHTING'), 1)
        weekly_water_h = round(weekly_schedule(schedules, 'WATER'), 1)
        schedules_cooling = schedules[['DAY', 'COOLING']].groupby('DAY').COOLING.value_counts()
        weekly_cooling_h = schedules_cooling['WEEKDAY'].SETPOINT * 5 + schedules_cooling['SATURDAY'].SETPOINT + \
                           schedules_cooling['SUNDAY'].SETPOINT

        # Extract relevant demand outputs
        Total_demand = pd.read_csv(locator.get_total_demand(), usecols=['GRID_MWhyr', 'GRID_a_MWhyr', 'GRID_l_MWhyr',
                                                                        'GRID_ve_MWhyr', 'GRID_ww_MWhyr',
                                                                        'GRID_cs_MWhyr',
                                                                        'GFA_m2', 'Af_m2', 'Aroof_m2', 'people0'])
        Annual_energy_demand_MWhyr = Total_demand.GRID_MWhyr.values[0]
        GFA_m2 = Total_demand.GFA_m2.values[0]
        EUI_kWhyr_m2 = Annual_energy_demand_MWhyr * 1000 / GFA_m2  # Energy Use Intensity
        EEI_kWhyr_m2 = EUI_kWhyr_m2 * 55 / weekly_occupancy_h  # Energy Efficiency Index, as defined by BCA, Singapore

        # Storage of outputs in a dict
        dict_outputs = {
            "instance": "ID" + str(index),
            "Annual_energy_demand_MWhyr": round(Annual_energy_demand_MWhyr, 2),
            "GFA_m2": round(GFA_m2, 2),
            "EUI_kWhyr_m2": round(EUI_kWhyr_m2, 2),
            "EEI_kWhyr_m2": round(EEI_kWhyr_m2, 2),
            "weekly_occupancy_h": weekly_occupancy_h,
            "weekly_appliances_h": weekly_appliances_h,
            "weekly_lighting_h": weekly_lighting_h,
            "weekly_water_h": weekly_water_h,
            "weekly_cooling_h": weekly_cooling_h
        }

        # Convert inputs and outputs into a csv file
        outputs_list.append(dict_outputs)
        outputs_df = pd.DataFrame(outputs_list)
        results = pd.concat([dataframe_with_instances, outputs_df], axis=1).reindex(dataframe_with_instances.index)
        results = pd.merge(dataframe_with_instances, outputs_df, on='instance', how='outer')
        # results.to_csv(locator.get_demand_results_folder() + r'\data_generation' + str(instance) + r'.csv', index=False)
        results.to_csv(locator.get_demand_results_folder() + r'\data_generation.csv', index=False)
    return results


def weekly_schedule(schedules, schedule_type='OCCUPANCY'):
    """
    This function sums the hourly fraction for a given schedule for one week.
    """
    schedule_hours = schedules[['DAY', schedule_type]].groupby('DAY').sum()[schedule_type]
    return schedule_hours['WEEKDAY'] * 5 + schedule_hours['SATURDAY'] + schedule_hours['SUNDAY']


def main(config):
    """
    The CLI will call this ``main`` function passing in a ``config`` object after adjusting the configuration
    to reflect parameters passed on the command line.

    :param config:
    :type config: cea.config.Configuration
    :return:
    """
    # Simulation general inputs
    number_simulations =config.scenario_generator.iterations
    print("Running for " + str(number_simulations) + " iterations")
    number_of_CPUs_to_keep_free = 1  # number cores that won't be used in this simulation (a minimum of 1 is indicated)
    number_cores_assigned = multiprocessing.cpu_count() - number_of_CPUs_to_keep_free

    # Prepare simulation scenarios and input data
    list_of_dataframe_with_instances = sampling_function(number_simulations, number_cores_assigned)  # prepare inputs
    locators_list = create_scenario(number_cores_assigned, config)  # prepare scenarios

    # Parallelization of simulations according to the number of cores available
    # data_generator_parallel = cea.utilities.parallel.vectorize(stochastic_scenario_generator, config.get_number_of_processes())
    data_generator_parallel = cea.utilities.parallel.vectorize(stochastic_scenario_generator, number_cores_assigned)
    n = len(locators_list)
    data_generator_parallel(
        repeat(config, n),
        locators_list,
        list_of_dataframe_with_instances)

    print("Simulation ended with " + str(number_simulations) + " iteration(s) split into " + str(
        number_cores_assigned) + " scenario(s)")


def sampling_function(number_simulations, number_cores):
    """
    This function creates a dataframe of random input data based on a distribution for each iteration.
    The inputs are split according the number of cores simulated, to facilitate parallelization.
    """

    # Creates input data for the simulations based on distributions.

    # Define distribution of all input variables. Beta distributions are used as default
    list_of_dict_variables = [{
        "instance": "ID" + str(i),
        "zone_floors_ag": round(beta.rvs(2, 2, scale=19, loc=2), 0),
        "zone_floor_to_floor_height": round(beta.rvs(8, 8, scale=2, loc=2), 1),
        "surrounding_floors_ag": round(beta.rvs(2, 2, scale=19, loc=1), 0),
        "surroundings_floor_to_floor_height": round(beta.rvs(8, 8, scale=2, loc=2), 1),
        "Hs_ag": round(beta.rvs(4, 4), 2),
        "Es": round(beta.rvs(8, 2), 2),  # TODO Es as independent variable, can make it conditional
        "Ns": round(beta.rvs(10, 2), 2),
        "void_deck": round(beta.rvs(3, 4, scale=2, loc=0), 0),
        "wwr": round(beta.rvs(12, 8), 2),
        "Cm_Af": round(beta.rvs(5, 7, scale=200, loc=100), 0) * 1000,
        "n50": round(beta.rvs(3, 7, scale=6, loc=1), 0),
        "U_win": round(beta.rvs(5, 6, scale=5, loc=0.5), 2),
        "G_win": round(beta.rvs(4, 4), 2),
        "e_win": round(beta.rvs(7, 4), 2),
        "F_F": round(beta.rvs(2, 8), 2),
        "U_roof": round(beta.rvs(2, 5, scale=0.85, loc=0.15), 2),
        "a_roof": round(beta.rvs(4, 4), 2),
        "e_roof": round(beta.rvs(8, 2), 2),
        "r_roof": round(beta.rvs(4, 4), 2),
        "U_wall": round(beta.rvs(5, 5, scale=3, loc=0.2), 2),
        "a_wall": round(beta.rvs(4, 4), 2),
        "e_wall": round(beta.rvs(10, 2), 2),
        "r_wall": round(beta.rvs(4, 6), 2),
        "rf_sh": round(beta.rvs(4, 6), 2),
        "Occ_m2p": round(beta.rvs(2, 4, scale=55, loc=3), 0),
        "Qs_Wp": round(beta.rvs(2, 10, scale=20, loc=70), 0),
        "X_ghp": round(beta.rvs(2, 10, scale=40, loc=80), 0),
        "Ea_Wm2": round(beta.rvs(5, 10, scale=15, loc=2), 1),
        "El_Wm2": round(beta.rvs(5, 10, scale=10, loc=5), 1),
        "Vww_ldp": round(beta.rvs(2, 2, scale=40, loc=0), 0),
        "Tcs_set_C": round(beta.rvs(3, 8, scale=5, loc=24), 1),
        "Ve_lsp": round(beta.rvs(2, 15, scale=20, loc=8), 1),
        "dT_Qcs": round(beta.rvs(3, 3, scale=1.6, loc=-2.5), 1),
        "ECONOMIZER": random.choice(['TRUE', 'FALSE']),
        "HEAT_REC": random.choice(['TRUE', 'FALSE']),
        "convection_cs": round(beta.rvs(10, 2), 1),
        "dTcs_C": round(beta.rvs(3, 3, scale=0.2, loc=0.5), 2),
        "efficiency_cooling": round(beta.rvs(3, 8, scale=4, loc=2.3), 2)
    } for i in range(number_simulations)]

    # The df is split according to the number of cores used in the simulation
    variables_df = pd.DataFrame(list_of_dict_variables)
    list_of_dataframe_with_instances = np.array_split(variables_df, number_cores)
    return list_of_dataframe_with_instances


def create_scenario(number_cores, config):
    """
    This function prepares the folders required for the simulation.
    Each scenario is assigned to one core, and overwritten across multiple iterations.
    """

    # Create folders to run simulations (one scenario per cpu) and saves list of locators
    base_scenario_path = config.scenario
    locators_list = []
    for core in range(number_cores):
        new_scenario_path = str(base_scenario_path) + str(core)
        shutil.rmtree(new_scenario_path, ignore_errors=True)
        shutil.copytree(base_scenario_path, new_scenario_path)
        locators_list.append(cea.inputlocator.InputLocator(new_scenario_path, config.plugins))
    return locators_list


if __name__ == '__main__':
    main(cea.config.Configuration())
